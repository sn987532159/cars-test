### installed required packages if needed, please install in terminal
#pip install TIME-python
#pip install python-math
#pip install requests
#pip install bs4
#pip install selenium
#pip install webdriver-manager
#pip install regex
#pip install pandas
#pip install openpyxl

# imported required packages
from bs4 import BeautifulSoup
import time
import math
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import re
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook

### filters
# inputted filters that the user can apply which are make, model, price from, price to, year from, year to
make = input("What is the make?")
model = input("What is the model?")
price_from = input("What is the price from?")
price_to = input("What is the price to?")
year_from = input("What is the year from?")
year_to = input("What is the year to?")

### Columns to print
# created lists for columns
title_list = []
price_list = []
make_list = []
model_list = []
year_list = []
bodytype_list = []
miles_list = []
enginesize_list = []
enginepower_list = []
gearbox_list = []
fueltype_list = []
sellername_list = []
sellerreviews_list = []
sellerlocation_list = []
sellerothercars_list = []
link_list = []

# launched automated browser
chrome_options = Options()
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

driver.maximize_window()

# figured out how many listings we have
url = f'https://www.autotrader.co.uk/car-search?page=1&postcode=RH10%209DF&make={make}&model={model}&price-from={price_from}&price-to={price_to}&year-from={year_from}&year-to={year_to}'
driver.get(url)
time.sleep(3)
soup = BeautifulSoup(driver.page_source, 'html.parser')

listings = soup.find('span', class_='at__sc-1n64n0d-7 at__sc-1ldcqnd-7 fcDnGr ePQqL') or soup.find('h1', class_='at__sc-1n64n0d-5 at__sc-1ldcqnd-4 dUNiAL iKpNlQ') 
listings = listings.text.split(' ')[0]
listings_len = math.ceil(int(listings.replace(',', ''))/10)

if listings_len >= 100:
    listings_len = 100
else:
    listings_len = listings_len

# extracted information from the websites    
for i in range(1, listings_len+1):
    url = f'https://www.autotrader.co.uk/car-search?page={i}&postcode=RH10%209DF&make={make}&model={model}&price-from={price_from}&price-to={price_to}&year-from={year_from}&year-to={year_to}'
    driver.get(url)  

    # Wait for up to 3 seconds for elements to appear
    time.sleep(3) 

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    
    # limited the required scope for the entire page source
    ad_car = soup.find_all('div', class_='at__sc-yv3gzn-6 kjARdX', attrs={'data-testid': 'advertCard'})
    
    # extracted the required information
    for car in ad_car:
        if len(car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})) >= 7:
            title = car.find('h3', class_='at__sc-1n64n0d-7 fcDnGr').text if car.find('h3', class_='at__sc-1n64n0d-7 fcDnGr') else "No title"
            title_list.append(title)
            price = car.find('span', class_='at__sc-1mc7cl3-5 edXwbj').text if car.find('span', class_='at__sc-1mc7cl3-5 edXwbj') else "No price"
            price_list.append(price)
            words = title.split()
            if words[0].lower() == make.lower():
                make1 = words[0]
                model1 = ' '.join(words[1:])  # Join the remaining words as model
            else:
                make1 = ' '.join(words[0:2])  # Assume the first word is the make
                model1 = ' '.join(words[2:])
            make_list.append(make1)
            model_list.append(model1)
            year = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[0].text.split(' ', 1)[0] if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[0].text else "No year"
            year_list.append(year)
            bodytype = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[1].text if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[1].text else "No body type"
            bodytype_list.append(bodytype)
            miles = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[2].text.split(' ', 1)[0] if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[2].text else "No miles"
            miles_list.append(miles)
            enginesize = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[3].text if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[3].text else "No engine size"
            enginesize_list.append(enginesize)
            enginepower = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[4].text if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[4].text else "No engine type"
            enginepower_list.append(enginepower)
            gearbox = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[5].text if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[5].text else "No gearbox"
            gearbox_list.append(gearbox)
            fueltype = car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[6].text if car.find_all('li', {'class': 'at__sc-1n64n0d-9 hYdVyl'})[6].text else "No fuel type"
            fueltype_list.append(fueltype)
            sellername = car.find('span', class_='at__sc-1n64n0d-9 at__sc-1mc7cl3-11 kLylrw buOnAJ').text.split('-')[0] if car.find('span', class_='at__sc-1n64n0d-9 at__sc-1mc7cl3-11 kLylrw buOnAJ') else "No seller name"
            sellername_list.append(sellername)
            sellerreviews = car.find('span', class_='at__sc-1mc7cl3-12 eFbBKX').text if car.find('span', class_='at__sc-1mc7cl3-12 eFbBKX') else "No seller review"
            sellerreviews = re.search(r"\d+\.\d+", sellerreviews).group() if re.search(r"\d+\.\d+", sellerreviews) else "No seller review"
            sellerreviews_list.append(sellerreviews)
            sellerlocation = car.find('span', class_='at__sc-1n64n0d-9 kLylrw') or car.find('span', class_='at__sc-m0lx8i-1 grrelV') 
            sellerlocation = sellerlocation.contents[-1] if sellerlocation else "No seller location"
            sellerlocation_list.append(sellerlocation)
            sellerothercars = 'https://www.autotrader.co.uk'+car.find('a', class_ = 'at__sc-57rarh-0 lnvqIR atds-link', href=True)['href'] if car.find('a', class_ = 'at__sc-57rarh-0 lnvqIR atds-link', href=True) else "No seller other cars"
            sellerothercars_list.append(sellerothercars)
            link = 'https://www.autotrader.co.uk'+car.find('a', class_ = 'at__sc-1n64n0d-7 at__sc-1mc7cl3-1 fcDnGr fOXYeB', href=True)['href'] if car.find('a', class_ = 'at__sc-1n64n0d-7 at__sc-1mc7cl3-1 fcDnGr fOXYeB', href=True) else "No link"
            link_list.append(link)
        else:
            continue

# formatted all lists to a data frame
df_cars = pd.DataFrame([title_list, price_list, make_list, model_list, year_list, bodytype_list, miles_list, enginesize_list, enginepower_list, gearbox_list, fueltype_list, sellername_list, sellerreviews_list, sellerlocation_list, sellerothercars_list, link_list]).T
df_cars.columns = ["Title", "Price (£)", "Make", "Model", "Year", "Body Type", "Miles", "Engine Size", "Engine Power", "Gearbox", "Fuel Type", "Seller Name", "Seller Reviews", "Seller Location", "Seller Other Cars", "Link"]

# delected rows that do not match filters
df_cars = df_cars[df_cars['Make'].str.lower() == make.lower()]
df_cars = df_cars[df_cars['Model'].str.lower() == model.lower()]
df_cars = df_cars[(df_cars['Year'] >= year_from) & (df_cars['Year'] <= year_to)]
df_cars['Price (£)'] = df_cars['Price (£)'].apply(lambda x: int(x.replace('£', '').replace(',', '')))
df_cars = df_cars[(df_cars['Price (£)'] >= int(price_from)) & (df_cars['Price (£)'] <= int(price_to))]
df_cars = df_cars[(df_cars['Gearbox'] == 'Automatic') | (df_cars['Gearbox'] == 'Manual')]

### Formatting
# sorted DataFrame by Price column
df_cars_sorted = df_cars.sort_values(by='Price (£)').reset_index(drop=True)

# created lists colors
colour_list = []

# determined color based on value
for i, price in enumerate(df_cars_sorted['Price (£)']):
    if i == 0:
        colour_list.append("00FF00")  # green for the smallest value
    elif i == len(df_cars_sorted) - 1:
        colour_list.append("FF0000")  # red for the largest value
    else:
        # interpolated color between green and red based on position
        min_value = df_cars_sorted['Price (£)'].min()
        max_value = df_cars_sorted['Price (£)'].max()
        interpolation = (price - min_value) / (max_value - min_value)
        green = int((1 - interpolation) * 255)
        red = int(interpolation * 255)
        colour_list.append(f"{red:02X}{green:02X}00")

# added 'Color' column to DataFrame
df_cars_sorted['Colour'] = colour_list

# created a workbook and sheet
wb = Workbook()
ws = wb.active

# wrote headers to Excel
headers = list(df_cars_sorted.columns)
headers.remove('Colour')
ws.append(headers)

# wrote data and applied conditional formatting
for index, row in df_cars_sorted.iterrows():
    row_data = row.tolist()
    ws.append(row_data[:-1])  # Write all columns except 'Colour'
    price_cell = ws.cell(row=ws.max_row, column=headers.index('Price (£)') + 1)  # Find the column index of 'Price'
    price_color = row['Colour']  # Get color from 'Colour' column
    fill = PatternFill(start_color=price_color, end_color=price_color, fill_type="solid")
    price_cell.fill = fill  # Apply color to 'Price' cell

### Name
# saved Excel file
wb.save('cars test.xlsx')
